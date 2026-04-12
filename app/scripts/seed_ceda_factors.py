import re
import uuid
import openpyxl
from decimal import Decimal, InvalidOperation
from pathlib import Path
from sqlalchemy.orm import Session
from app.database import engine
from app.models.emission_factors import EmissionFactor
from app.models.user import User
from app.models.category import Category

PROVIDER = "Open CEDA"
UNIT_OF_MEASURE = "USD"
YEAR = 2023
VERSION = "2025-11-12"

def get_system_user(session: Session) -> uuid.UUID:
    sys_email = "system_epa@scopeops.local"
    sys_user = session.query(User).filter(User.email == sys_email).first()
    if not sys_user:
        sys_user = User(email=sys_email, full_name="EPA System Administrator", provider="system", is_active=True)
        session.add(sys_user)
        session.commit()
        session.refresh(sys_user)
    return sys_user.id

def parse_decimal(value: object) -> Decimal | None:
    if value is None: return None
    text = str(value).strip()
    if not text or text.lower() in {"na", "n/a", "nan", "null"}: return None
    cleaned = text.replace(",", "")
    if cleaned.startswith("(") and cleaned.endswith(")"): cleaned = f"-{cleaned[1:-1]}"
    try:
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None

def format_external_id(sector_code: str) -> str:
    slug = sector_code.strip().upper()
    return f"OPEN-CEDA-2025-{slug}"

def seed_ceda_factors(xlsx_path: Path) -> None:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    
    sheet_conv = next((s for s in wb.sheetnames if "conversion" in s.lower()), None)
    ws_conv = wb[sheet_conv]
    
    code_row = []
    mult_row = []
    
    for row in ws_conv.iter_rows(values_only=True):
        row_strs = [str(c).strip() if c is not None else "" for c in row]
        row_lower = [x.lower() for x in row_strs]
        
        if any("1111a0" in x for x in row_lower):
            code_row = row_strs
        elif any("purchaser" in x and "producer" in x for x in row_lower) and len([x for x in row_strs if x]) > 5:
            mult_row = row_strs
            
        if code_row and mult_row:
            break
            
    conversions = {}
    for code, mult_val in zip(code_row, mult_row):
        code_str = code.strip()
        if not code_str or code_str.lower() in ["sector name", "sector code", "purchaser - producer conversion", "source"]:
            continue
            
        mult = parse_decimal(mult_val)
        if mult:
            conversions[code_str] = mult

    print(f"Extracted {len(conversions)} conversion multipliers.")


    #  Process Raw Factors (Horizontal Matrix)
    
    sheet_raw = next((s for s in wb.sheetnames if "raw" in s.lower()), None)
    ws_raw = wb[sheet_raw]
    
    codes_row = []
    data_rows = []
    found_headers = False
    
    for row in ws_raw.iter_rows(values_only=True):
        row_strs = [str(c).strip() if c is not None else "" for c in row]
        row_lower = [x.lower() for x in row_strs]
        
        if not found_headers:
            if any("1111a0" in x for x in row_lower):
                codes_row = row_strs
                found_headers = True
        else:
            if any(row_strs):
                data_rows.append(row_strs)
                
    print(f"Detected {len(data_rows)} Country Rows to process.")
    wb.close()


    # FAST In-Memory Data Prep & Bulk Insert

    new_factors_to_insert = []

    # Open a brief session just to grab existing data and categories
    with Session(engine) as session:
        sys_user_id = get_system_user(session)
        category_map = {c.category_id: c.category_name for c in session.query(Category).all()}
        
        print("Downloading existing database index to memory (preventing duplicates)...")
        # Grab only the identifiers to keep memory usage tiny
        existing_tuples = session.query(EmissionFactor.external_id, EmissionFactor.geography).all()
        existing_set = set(existing_tuples) # O(1) lookup speed

    print("Building global carbon matrix in memory...")
    for row in data_rows:
        country_name = str(row[1]).strip() 
        
        if not country_name or country_name.lower() in ["country", "geography", ""]:
            continue
            
        for col_idx in range(len(row)):
            if col_idx >= len(codes_row):
                continue
                
            code = codes_row[col_idx].strip()
            
            if not code or code.lower() in ["country code", "country", "country ", "geography"]:
                continue
                
            base_ef = parse_decimal(row[col_idx])
            multiplier = conversions.get(code)
            
            if not base_ef or not multiplier:
                continue

            final_ef = base_ef * multiplier
            ext_id = format_external_id(code)

            # INSTANT Memory Check instead of slow Database Query
            if (ext_id, country_name) not in existing_set:
                sector_name = category_map.get(code, f"Sector {code}")
                
                new_factors_to_insert.append(EmissionFactor(
                    external_id=ext_id, provider=PROVIDER, name=f"{sector_name} ({code})",
                    geography=country_name, year=YEAR, unit_of_measure=UNIT_OF_MEASURE,
                    co2e_per_unit=final_ef, scope_3_intensity=final_ef,
                    owner_id=sys_user_id, version=VERSION,
                    methodology=f"Open CEDA 2025 Global Factors - {country_name}"
                ))
                # Add to set so we don't accidentally queue duplicates if the Excel file has them
                existing_set.add((ext_id, country_name))

    print(f"Prepared {len(new_factors_to_insert)} new global factors. Blasting to database...")

    # Open fresh sessions to bulk save blocks of data, preventing connection timeouts
    batch_size = 5000
    inserted_count = 0
    
    for i in range(0, len(new_factors_to_insert), batch_size):
        batch = new_factors_to_insert[i:i + batch_size]
        # Using a fresh session for each batch ensures the connection stays alive
        with Session(engine) as session:
            session.add_all(batch)
            session.commit()
            inserted_count += len(batch)
            print(f"   ...successfully committed {inserted_count} / {len(new_factors_to_insert)}...")

    print(f"Master seed complete! {inserted_count} Global Open CEDA factors added.")

if __name__ == "__main__":
    default_xlsx = Path(__file__).resolve().parent.parent.parent / "data" / "Open CEDA 2025.xlsx"
    seed_ceda_factors(default_xlsx)