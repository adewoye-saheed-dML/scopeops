import uuid
import openpyxl
from decimal import Decimal
from sqlalchemy.orm import Session
from app.database import engine
from app.models.emission_factors import EmissionFactor
from app.models.user import User

# Global Warming Potentials from Table 11 of the EPA data
GWP_CH4 = 28
GWP_N2O = 265
LBS_TO_KG = 0.453592

def get_system_user(session: Session) -> uuid.UUID:
    sys_email = "system_epa@scopeops.local"
    sys_user = session.query(User).filter(User.email == sys_email).first()
    if not sys_user:
        sys_user = User(email=sys_email, full_name="EPA System Administrator", provider="system", is_active=True)
        session.add(sys_user)
        session.commit()
        session.refresh(sys_user)
    return sys_user.id

def parse_epa_excel(file_path: str):
    factors = []
    current_table = None

    print(f"Loading Excel file: {file_path} (This may take a few seconds...)")
    # data_only=True ensures we read the calculated values, not the Excel formulas
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # Intelligently find the main data sheet, or fallback to the active one
    sheet_name = None
    for name in wb.sheetnames:
        if "Emission" in name or "Factors" in name:
            sheet_name = name
            break
    sheet = wb[sheet_name] if sheet_name else wb.active
    
    for row in sheet.iter_rows(values_only=True):
        # Convert cells to string, replace None with ""
        r = [str(cell).strip() if cell is not None else "" for cell in row]
        
        # Combine row into a single string for easy searching
        row_text = " ".join(r).strip()
        
        if not row_text:
            continue
            
        # --- Detect which table we are currently reading ---
        if "Table 6" in row_text:
            current_table = "electricity"
            continue
        elif "Table 9" in row_text:
            current_table = "waste"
            continue
        elif "Table " in row_text:
            current_table = None # Skip other tables
            continue
            
        # --- PARSE TABLE 6: ELECTRICITY (eGRID Subregions) ---
        if current_table == "electricity":
            if "Source:" in row_text:
                current_table = None
                continue
            
            # Because of Excel formatting, the acronym is pushed to index 2
            if len(r) > 6:
                acronym = r[2]
                if len(acronym) == 4 or acronym == "US Average":
                    try:
                        name = r[3]
                        co2_lb = float(r[4])
                        ch4_lb = float(r[5])
                        n2o_lb = float(r[6])
                        
                        total_co2e_lbs = co2_lb + (ch4_lb * GWP_CH4) + (n2o_lb * GWP_N2O)
                        co2e_kg = total_co2e_lbs * LBS_TO_KG
                        
                        factors.append({
                            "external_id": f"EPA-EGRID-2025-{acronym}",
                            "provider": "EPA eGRID",
                            "name": f"Purchased Electricity - {name}",
                            "geography": acronym,
                            "year": 2025,
                            "unit_of_measure": "MWh", 
                            "co2e_per_unit": co2e_kg,
                            "scope_1_intensity": 0.0,
                            "scope_2_intensity": co2e_kg, 
                            "scope_3_intensity": 0.0,
                            "source_url": "https://www.epa.gov/climateleadership/ghg-emission-factors-hub",
                            "methodology": "eGRID Total Output Emission Rates",
                            "version": "2025.1"
                        })
                    except (ValueError, IndexError):
                        continue 
        
        # --- PARSE TABLE 9: WASTE (Scope 3 Category 5) ---
        elif current_table == "waste":
            if "Source:" in row_text:
                current_table = None
                continue
            
            if len(r) > 4:
                material_name = r[2]
                
                # Check if it's a valid material row
                if material_name and material_name not in ["Material", "Metric Tons CO2e / Short Ton Material"]:
                    # Because of formatting, landfilled is index 4
                    landfilled_val = r[4] 
                    
                    if landfilled_val and landfilled_val not in ["NA", ""]:
                        try:
                            mt_per_short_ton = float(landfilled_val)
                            kg_per_kg = (mt_per_short_ton * 1000) / 907.185
                            
                            factors.append({
                                "external_id": f"EPA-WASTE-LF-{material_name.replace(' ', '').upper()}",
                                "provider": "EPA WARM",
                                "name": f"Waste Disposal (Landfilled) - {material_name}",
                                "geography": "US",
                                "year": 2025,
                                "unit_of_measure": "kg",
                                "co2e_per_unit": kg_per_kg,
                                "scope_1_intensity": 0.0,
                                "scope_2_intensity": 0.0,
                                "scope_3_intensity": kg_per_kg, 
                                "source_url": "https://www.epa.gov/climateleadership/ghg-emission-factors-hub",
                                "methodology": "WARM Landfilling Emissions",
                                "version": "2025.1"
                            })
                        except (ValueError, IndexError):
                            continue

    return factors

def seed_epa_factors():
    print("Starting EPA Emission Factor Seeder...")
    
    excel_file_path = "data/ghg-emission-factors-hub-2025.xlsx" 
    
    parsed_factors = parse_epa_excel(excel_file_path)
    print(f"Extracted {len(parsed_factors)} factors from the Excel file.")
    
    with Session(engine) as session:
        sys_user_id = get_system_user(session)
        added_count = 0

        for factor_data in parsed_factors:
            # Idempotency check: Don't duplicate if we already seeded it
            exists = session.query(EmissionFactor).filter_by(
                external_id=factor_data["external_id"],
                year=factor_data["year"]
            ).first()

            if not exists:
                new_factor = EmissionFactor(
                    **factor_data,
                    owner_id=sys_user_id
                )
                session.add(new_factor)
                added_count += 1

        session.commit()
        print(f"Successfully seeded {added_count} new EPA factors into the database.")

if __name__ == "__main__":
    seed_epa_factors()